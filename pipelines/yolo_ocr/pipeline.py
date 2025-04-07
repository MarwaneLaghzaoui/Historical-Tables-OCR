import cv2
import numpy as np
import matplotlib.pyplot as plt
import fitz 
import os
from PIL import Image
import pytesseract
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
import shutil
from PyPDF2 import PdfReader, PdfWriter
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from correct_skew_angle import correct_pdf_skew_angle


global colonne_noire_positions
colonne_noire_positions = []
global ligne_noire_positions
ligne_noire_positions = []

global lignes_noires_colonne_gauche_save
lignes_noires_colonne_gauche_save = []

global lignes_noires_inserees
lignes_noires_inserees = []


Tk().withdraw()  # Ne pas afficher la fenêtre principale Tk
pdf_path = askopenfilename(title="Choisir un fichier PDF", filetypes=[("PDF files", "*.pdf")])

file_name = os.path.splitext(os.path.basename(pdf_path))[0]
pdf_file_name = os.path.basename(pdf_path)
pdf_folder = os.path.dirname(pdf_path) + os.sep

correct_pdf_skew_angle(pdf_folder,pdf_file_name)

pdf_straightened_path = os.path.join(pdf_folder, "output_" + pdf_file_name)
pdf_document = fitz.open(pdf_straightened_path)



#Coefficients à ajuster

min_area_points = 10
max_area_points = 60
distance_max_point = 5
tolerance_apres_ligne = 5

terme_longueur_ajoutee = 20



def manual_threshold(image, threshold):
    thresholded_image = np.zeros_like(image)
    thresholded_image[image > threshold] = 255
    return thresholded_image

def detection_premiere_colonne_noire(binary_image, colored_image):
    seuil_noir = 0.03
    hauteur, largeur = binary_image.shape
    x = 0
    while x < largeur:
        colonne = binary_image[:, x]
        noirs = np.sum(colonne == 0)
        if noirs / hauteur >= seuil_noir:
            colored_image[:, x, :] = [255, 0, 0]
            if x >= largeur:
                break
            colonne = binary_image[:, x]
            noirs = np.sum(colonne == 0)
            colonne_noire_positions.append(x)
            x = largeur + 1
        x += 1

def detection_colonnes_noires(binary_image, longueur_min, espace_min, tolerance, decalage_premiere_colonne,colored_image):
    hauteur, largeur = binary_image.shape
    
    
    x = colonne_noire_positions[0] + decalage_premiere_colonne
    while x < largeur:
        colonne = binary_image[:, x]
        compteur_noirs = 0
        compteur_trous = 0
        debut_colonne = None

        # Vérifier si la colonne contient une séquence noire assez longue
        for y in range(hauteur):
            if colonne[y] == 0:  # Pixel noir
                if compteur_noirs == 0:
                    debut_colonne = y
                compteur_noirs += 1
                compteur_trous = 0  # Réinitialiser le compteur de trous
            else:
                if compteur_noirs > 0 and compteur_trous < tolerance:
                    compteur_trous += 1  # Tolérer un petit trou
                else:
                    if compteur_noirs >= longueur_min:  # Colonne noire validée
                        if not colonne_noire_positions or (x - colonne_noire_positions[-1] >= espace_min):
                            colonne_noire_positions.append(x)
                            colored_image[:, x, :] = [255, 0, 0]  # Colorer en rouge
                            x += espace_min -1 # Avancer mais pas trop loin
                    compteur_noirs = 0
                    compteur_trous = 0  # Réinitialiser

        # Vérification si la colonne se termine par une séquence noire valide
        if compteur_noirs >= longueur_min:
            if not colonne_noire_positions or (x - colonne_noire_positions[-1] >= espace_min):
                colonne_noire_positions.append(x)
                colored_image[:, x, :] = [255, 0, 0]  # Colorer en rouge
                x += espace_min - 1  # Avancer mais sans sauter trop loin

        x += 1  # On avance toujours d’au moins un pixel

    return colonne_noire_positions

def detection_lignes_noires(binary_image, longueur_min, espace_min, tolerance,colored_image):
    hauteur, largeur = binary_image.shape
    
    y = 0
    while y < hauteur:
        ligne = binary_image[y, :]
        compteur_noirs = 0
        compteur_trous = 0
        debut_ligne = None

        # Vérifier si la ligne contient une séquence noire assez longue
        for x in range(colonne_noire_positions[1]):
            if ligne[x] == 0:  # Pixel noir
                if compteur_noirs == 0:
                    debut_ligne = x
                compteur_noirs += 1
                compteur_trous = 0  # Réinitialiser le compteur de trous
            else:
                if compteur_noirs > 0 and compteur_trous < tolerance:
                    compteur_trous += 1  # Tolérer un petit trou
                else:
                    if compteur_noirs >= longueur_min:  # Ligne noire validée
                        if not ligne_noire_positions or (y - ligne_noire_positions[-1] >= espace_min):
                            ligne_noire_positions.append(y)
                            colored_image[y, :, :] = [0, 0, 255]  # Colorer en bleu
                            y += espace_min -1 # Avancer mais pas trop loin
                    compteur_noirs = 0
                    compteur_trous = 0  # Réinitialiser

        # Vérification si la ligne se termine par une séquence noire valide
        if compteur_noirs >= longueur_min:
            if not ligne_noire_positions or (y - ligne_noire_positions[-1] >= espace_min):
                ligne_noire_positions.append(y)
                colored_image[y, :, :] = [0, 0, 255]  # Colorer en bleu
                y += espace_min - 1  # Avancer mais sans sauter trop loin

        y += 1  # On avance toujours d’au moins un pixel

    return ligne_noire_positions


# Fonction pour insérer des lignes avec décalage
def inserer_lignes_decalees(binary_image, start_col, end_col, lignes_noires_colonne_gauche,coefficient_haut):
    hauteur, largeur = binary_image.shape
    lignes_blanches = np.full((int(3*coefficient_haut), largeur), 255, dtype=np.uint8)
    lignes_noires = np.full((int(2*coefficient_haut), largeur), 255, dtype=np.uint8)
    lignes_noires[:, start_col:end_col] = 1
    sequence_lignes = np.vstack((lignes_blanches, lignes_noires, lignes_blanches))
    
    image_etendue = binary_image.copy()
    decalage = 0
    for y in sorted(lignes_noires_colonne_gauche):
        if y + decalage < image_etendue.shape[0]:
            image_etendue = np.insert(image_etendue, y + int(4*coefficient_haut) + decalage, sequence_lignes, axis=0)
            decalage += 2*int(3*coefficient_haut)+int(2*coefficient_haut)
    return image_etendue


def recadrer_cellule(x,y):
    
    dossier = os.path.join(os.path.dirname(__file__), "cellules")
    dossier_recadre = os.path.join(os.path.dirname(__file__), "cellules recadrées")
    
    image_path = os.path.join(dossier, f'cell_{x}_{y}.png')
    
    image = Image.open(image_path).convert("L")


    width, height = image.size
    pixels = image.load()


    new_left = int(0.2 * width)
    new_right = int(0.99 * width)
    new_top = int(0.25 * height)
    crop_box = (new_left, new_top, new_right, height)
    cropped_image = image.crop(crop_box)

    cropped_width, cropped_height = cropped_image.size
    cropped_pixels = cropped_image.load()


    coord_premier_noir_hauteur = None
    coord_premier_noir_largeur = None


    for i in range(cropped_height):
        for j in range(cropped_width):
            if cropped_pixels[j, i] < 10: 
                coord_premier_noir_hauteur = i
                break
        if coord_premier_noir_hauteur is not None:
            break

    for j in range(cropped_width):
        for i in range(cropped_height):
            if cropped_pixels[j, i] < 10:
                coord_premier_noir_largeur = j
                break
        if coord_premier_noir_largeur is not None:
            break

    if coord_premier_noir_hauteur is None:
        coord_premier_noir_hauteur = 0
    if coord_premier_noir_largeur is None:
        coord_premier_noir_largeur = 0

    coord_premier_noir_hauteur = max(coord_premier_noir_hauteur - 10, 0)
    coord_premier_noir_largeur = max(coord_premier_noir_largeur - 10, 0)

    crop_box = (
        coord_premier_noir_largeur,
        coord_premier_noir_hauteur,
        cropped_width,
        cropped_height
    )

    final_image = cropped_image.crop(crop_box)

    cell_path = os.path.join(dossier_recadre, f'cell_{x}_{y}.png')
    final_image.save(cell_path)


def traitement_page(indice):

    
    page = pdf_document[indice]
    pix = page.get_pixmap(dpi=100)
    image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
    image_gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    
    binary_image = manual_threshold(image_gray, 200)
    
    
    #calcul du coefficient de résolution
    hauteur, largeur = binary_image.shape
    coefficient = hauteur * largeur / (3450 * 2659)
    coefficient_haut = hauteur / 3450 + 1
    coefficient_larg = largeur / 2659 + 1
    
    # Détection des colonnes noires
    
    colored_image = np.stack([binary_image] * 3, axis=-1)
    
    detection_premiere_colonne_noire(binary_image,colored_image)
    
    detection_colonnes_noires(binary_image, 200* int(coefficient_haut), 40* int(coefficient_haut),5* int(coefficient_haut), 40* int(coefficient_larg), colored_image)
    
    detection_lignes_noires(binary_image, 200* int(coefficient_larg), 40 * int(coefficient_larg), 0* int(coefficient_larg), colored_image)
    
    
    if(colonne_noire_positions[1]- colonne_noire_positions[0] > 200 * coefficient_larg) :
        
        print("Traitement de la page " + str(indice) + " (recto)")
        lignes_noires_colonne_gauche_save = []

        #On fait face au recto (distance premiere colonne a deuxieme superireure a 300 pixels)
        
        # Définir la zone d'intérêt entre les colonnes
        zone_start = colonne_noire_positions[0]
        zone_end = colonne_noire_positions[1] + int(tolerance_apres_ligne * coefficient_larg)
        ligne_depart = ligne_noire_positions[1]
        zone_interet = binary_image[ligne_noire_positions[1]:, zone_start:zone_end]

        # Sauvegarder la zone d'intérêt
        zone_interet_sauvegardee = zone_interet.copy()

        # Détection des points dans la zone d'intérêt
        min_area = min_area_points * coefficient
        max_area = max_area_points * coefficient

        contours, _ = cv2.findContours(255 - zone_interet, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        binary_image_colored = np.stack([binary_image] * 3, axis=-1)
        droite_points = {}

        for contour in contours:
            area = cv2.contourArea(contour)
            if min_area <= area <= max_area:
                x, y, w, h = cv2.boundingRect(contour)
                if (x > 0 and y > 0 and x + w < zone_interet.shape[1] and y + h < zone_interet.shape[0]):
                    surrounding_area = zone_interet[y-1:y+h+1, x-1:x+w+1]
                    if np.all(surrounding_area[0, :] == 255) and np.all(surrounding_area[-1, :] == 255) and \
                       np.all(surrounding_area[:, 0] == 255) and np.all(surrounding_area[:, -1] == 255):
                        cv2.rectangle(colored_image, (x + zone_start, y + ligne_depart ), (x + w + zone_start, y + ligne_depart + h), (0, 0, 255), 2)
                        ligne = y // 10
                        if (ligne not in droite_points or x + w > droite_points[ligne][0]):
                            droite_points[ligne] = (x + w, y + ligne_depart)

        droite_colonne = zone_interet.shape[1] - 1

        filtered_droite_points = {ligne: point for ligne, point in droite_points.items()
                                  if abs(droite_colonne - point[0]) <= (int(distance_max_point * coefficient_larg) + int(tolerance_apres_ligne * coefficient_larg))}

        # Dessiner des rectangles distincts pour les points restants et ajouter un numéro
        for i, (ligne, point) in enumerate(filtered_droite_points.items(), start=1):
            cv2.rectangle(colored_image, (point[0] + zone_start, point[1]),
                          (point[0] + zone_start + 5, point[1] + 5), (0, 255, 0), 2)
            cv2.putText(colored_image, str(i), (point[0] + zone_start + 7, point[1] + 5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1, cv2.LINE_AA)

        last_ligne, last_point = list(filtered_droite_points.items())[-1]
        last_y = last_point[1]
        new_ligne_depart = last_y - 10
        new_zone_interet = binary_image[new_ligne_depart:, zone_start:zone_end]
        new_zone_interet_sauvegardee = new_zone_interet.copy()


        lignes_noires_colonne_gauche = [point[1] for point in filtered_droite_points.values()]

        
        lignes_noires_colonne_gauche_save = [x - ligne_noire_positions[0] for x in lignes_noires_colonne_gauche]

        # Affichage des détections
        '''plt.figure(figsize=(10, 10))
        plt.imshow(colored_image)
        plt.axis('off')
        plt.show()'''

        # Recherche de la première colonne blanche après la deuxième colonne noire
        colonne_blanche = None
        for i in range(colonne_noire_positions[1] + 1, binary_image.shape[1]):
            if np.all(binary_image[:, i] == 255):  # Colonne composée uniquement de blancs
                colonne_blanche = i
                break            

        image_etendue = inserer_lignes_decalees(binary_image, colonne_noire_positions[1], colonne_blanche, lignes_noires_colonne_gauche,coefficient_haut)

        # Étirement de la zone sauvegardée
        nouvelle_hauteur = image_etendue.shape[0] - ligne_noire_positions[1] + int(terme_longueur_ajoutee*coefficient_haut)
        zone_interet_etiree = cv2.resize(new_zone_interet_sauvegardee, (new_zone_interet.shape[1], nouvelle_hauteur), interpolation=cv2.INTER_LINEAR)
        hauteur_disponible = image_etendue.shape[0] - new_ligne_depart
        zone_interet_etiree = zone_interet_etiree[:hauteur_disponible, :]

        # Réinsérer la zone étirée
        image_etendue[new_ligne_depart:, zone_start:zone_end] = zone_interet_etiree

        # Sauvegarde de l'image finale        
        plt.figure(figsize=(10, 10))
        plt.imshow(image_etendue, cmap='gray')
        plt.axis('off')
        #plt.show()
        plt.imsave(f"pipelines/yolo_ocr/pages traitées/{indice}.pdf", image_etendue, cmap='gray')
        
        '''img_pil = Image.fromarray(image_etendue)
        img_path = os.path.join(dossier_page, f'{indice}.pdf')
        img_pil.save(img_path)'''
    
    else :
        
        '''plt.figure(figsize=(10, 10))
        plt.imshow(colored_image)
        plt.axis('off')
        plt.show()'''
        
        #On fait face au verso        
        print("Traitement de la page " + str(indice) + " (verso)")
        
        colonne_blanche = None
        for i in range(colonne_noire_positions[1] + 1, binary_image.shape[1]):
            if np.all(binary_image[:, i] == 255):  # Colonne composée uniquement de blancs
                colonne_blanche = i
                break            
        lignes_noires_colonne_gauche = [x + ligne_noire_positions[0] for x in lignes_noires_colonne_gauche_save]        
        image_etendue = inserer_lignes_decalees(binary_image, colonne_noire_positions[0], colonne_blanche, lignes_noires_colonne_gauche,coefficient_haut)
        
        # Sauvegarde de l'image finale
        plt.figure(figsize=(10, 10))
        plt.imshow(image_etendue, cmap='gray')
        plt.axis('off')
        #plt.show()        
        plt.imsave(f"pipelines/yolo_ocr/pages traitées/{indice}.pdf", image_etendue, cmap='gray')


def lecture_ocr(indice):
    
    pdf_path = f"pipelines/yolo_ocr/pages traitées/{indice}.pdf"
    pdf_document = fitz.open(pdf_path)
    
    page = pdf_document[0]
    pix = page.get_pixmap(dpi=100)
    image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
    binary_image = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    
    hauteur, largeur = binary_image.shape
    colored_image = np.stack([binary_image] * 3, axis=-1)  # Convertir en RGB
    
    y = 0
    while y < hauteur:
        point_milieu = binary_image[y, largeur//2]
        if point_milieu == 1 :  # Pixel noir
            lignes_noires_inserees.append(y)
            y+=4
        y+=1
    
    taille_ligne = []
    
    for i in range(len(lignes_noires_inserees)-1):
        taille_ligne.append(abs(lignes_noires_inserees[i]-lignes_noires_inserees[i+1]))
    lignes_noires_inserees.insert(0,(lignes_noires_inserees[0]-np.median(taille_ligne)))
    
    #print(lignes_noires_inserees)
    
    taille_colonne = []
    
    for i in range(len(colonne_noire_positions)-2):
        taille_colonne.append(abs(colonne_noire_positions[i+1]-colonne_noire_positions[i+2]))
    colonne_noire_positions.append(int(colonne_noire_positions[len(colonne_noire_positions)-1]+np.median(taille_colonne)))
    
    #print(colonne_noire_positions[1:])
    
    for y in lignes_noires_inserees:
        i = int(y)
        colored_image[i, :, :] = [255, 0, 0]  # Colorer en rouge
        
    for y in colonne_noire_positions[1:]:
        i = int(y)
        colored_image[:, i, :] = [255, 0, 0]  # Colorer en rouge
    
    '''plt.axis('off')
    plt.imshow(colored_image)
    plt.show()'''
    
    
    dossier = os.path.join(os.path.dirname(__file__), "cellules")
    if os.path.exists(dossier):
        shutil.rmtree(dossier)
    os.makedirs(dossier, exist_ok=True)

    data = {}
    
    config = f"--psm 7 --oem 3 -l digitdetector  -c tessedit_char_whitelist=0123456789"
    
    # Sauvegarde des cellules extraites et OCR
    for i in range(len(lignes_noires_inserees) - 1):
        for j in range(len(colonne_noire_positions) - 2):
            y1, y2 = int(lignes_noires_inserees[i]), int(lignes_noires_inserees[i+1])
            x1, x2 = int(colonne_noire_positions[j+1]), int(colonne_noire_positions[j+2])
            cellule = image[y1:y2, x1:x2]
            
            # Sauvegarde de l'image
            img_pil = Image.fromarray(cellule)
            cell_path = os.path.join(dossier, f'cell_{i}_{j}.png')
            img_pil.save(cell_path)
     
    dossier_recadre = os.path.join(os.path.dirname(__file__), "cellules recadrées")
    if os.path.exists(dossier_recadre):
        shutil.rmtree(dossier_recadre)
    os.makedirs(dossier_recadre, exist_ok=True)
    
    for x in range(len(lignes_noires_inserees) - 1):
        for y in range(len(colonne_noire_positions) - 2):
            recadrer_cellule(x,y)
        
    # Nombre total d'itérations pour tqdm
    total = (len(lignes_noires_inserees) - 1) * (len(colonne_noire_positions) - 2)

    # Barre de progression
    with tqdm(total=total, desc="Traitement OCR", unit="cellule") as pbar:
        for i in range(len(lignes_noires_inserees) - 1):
            for j in range(len(colonne_noire_positions) - 2):
                
                # OCR avec tesseract
                cell_path = os.path.join(dossier_recadre, f'cell_{i}_{j}.png')
                text = pytesseract.image_to_string(Image.open(cell_path), config=config).strip()

                # Stocker dans un dictionnaire pour correspondre aux coordonnées Excel
                if i not in data:
                    data[i] = {}
                data[i][j] = text
                
                # Mise à jour de la barre de progression
                pbar.update(1)

        # Création du fichier Excel avec les données bien placées
        excel_path = os.path.join(os.path.dirname(__file__), "csv_tables/"+file_name+".xlsx")
        
        df = pd.DataFrame.from_dict(data, orient='index').sort_index()
        
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=f"Page {indice}", index=True, header=True)
        

        wb = load_workbook(excel_path)
        if "Temp" in wb.sheetnames:
            wb.remove(wb["Temp"])
        wb.save(excel_path)
        wb.close()
        
        print(f"Cellules extraites et sauvegardées dans {dossier}")
        print(f"Résultats OCR enregistrés dans {excel_path}")


def traitement_document(nombre_pages):
    
    global colonne_noire_positions, ligne_noire_positions, lignes_noires_inserees

    BASE_DIR = os.path.dirname(__file__)

    dossier_page = os.path.join(BASE_DIR, "pages traitées")
    if os.path.exists(dossier_page):
        shutil.rmtree(dossier_page)
    os.makedirs(dossier_page, exist_ok=True)

    excel_path = os.path.join(BASE_DIR, "csv_tables/"+file_name+".xlsx")
    if os.path.exists(excel_path):
        os.remove(excel_path)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_temp = pd.DataFrame({"Temp": ["Feuille temporaire"]})
        df_temp.to_excel(writer, sheet_name="Temp", index=False)

    for p in range(nombre_pages):
        traitement_page(p)
        lecture_ocr(p)
        colonne_noire_positions = []
        ligne_noire_positions = []
        lignes_noires_inserees = []
        
    # Création d'un objet PdfWriter pour le fichier final
    pdf_writer = PdfWriter()

    for i in range(nombre_pages):
        fichier_pdf = f"pipelines/yolo_ocr/pages traitées/{i}.pdf"
        try:
            pdf_reader = PdfReader(fichier_pdf)
            # Ajout de chaque page du PDF courant dans le fichier final
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)
        except FileNotFoundError:
            print(f"Le fichier {fichier_pdf} est introuvable.")
        except Exception as e:
            print(f"Erreur avec {fichier_pdf} : {e}")

    with open("pages_traitees.pdf", "wb") as fichier_sortie:
        pdf_writer.write(fichier_sortie)

    print("Toutes les pages ont été traitées.")




traitement_document(2)





